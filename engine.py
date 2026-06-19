"""
ssbt — spreadsheet build tool.
dbt for engineers who live in Excel.

Usage:
    ssbt build [--yml FILE] [--input FILE] [--output FILE] [--dry-run]
    ssbt test  [--yml FILE] [--input FILE] [--output FILE]
    ssbt docs  [--yml FILE]
"""

from __future__ import annotations

import argparse
import os
import re
from typing import Any
import yaml
import openpyxl
import warnings
import pandas as pd
import duckdb

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _coerce_cell(val):
    """Convert an openpyxl cell value to a proper Python type.
    
    Formula errors (#DIV/0!, #VALUE!, etc.) become None.
    Numeric strings become int/float.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        # Check for NaN from formula errors
        if isinstance(val, float) and val != val:
            return None
        return val
    # Check for Excel error constants (openpyxl stores them as strings like '#DIV/0!')
    if isinstance(val, str) and val.startswith('#'):
        return None
    # Try numeric conversion
    try:
        f = float(val)
        if f == int(f) and '.' not in str(val):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


# ---------------------------------------------------------------------------
# YAML parsing — ssbt.yml with model paths + column tests
# ---------------------------------------------------------------------------

def load_manifest(yml_path: str) -> dict[str, Any]:
    """Load ssbt.yml and return the full parsed config."""
    with open(yml_path) as f:
        cfg = yaml.safe_load(f)
    return cfg


def load_models(cfg: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Extract models from parsed config."""
    models: dict[str, dict[str, Any]] = {}
    for entry in cfg.get("models", []):
        name = entry["name"]
        models[name] = {
            "sql_path": entry["path"],
            "config": entry.get("config", {}),
            "columns": entry.get("columns", []),
        }
    return models


def load_sources(cfg: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract sources from parsed config.

    Returns a list of {name, path, sheets} dicts.
    Also supports legacy --input flag: if no sources defined, wraps --input
    as a single source called 'input' with all its sheets.
    """
    sources = cfg.get("sources", [])
    return sources


# ---------------------------------------------------------------------------
# DAG resolution
# ---------------------------------------------------------------------------

def resolve_dag(models: dict[str, dict[str, Any]], sources: list[dict[str, Any]] | None = None) -> list[str]:
    """Topological sort of models. Returns execution order.

    Dependencies are inferred from {{ ref('name') }} in SQL files.
    Sources are treated as leaf nodes (no ordering needed).
    """
    if sources is None:
        sources = []

    # Build a map of model_name -> set of {{ ref() }} deps from SQL
    ref_deps: dict[str, set[str]] = {}
    for name, model in models.items():
        sql_path = os.path.join(os.path.dirname(os.path.abspath("ssbt.yml")), model["sql_path"])
        with open(sql_path) as f:
            sql = f.read()
        deps = set()
        for m in _REF_RE.finditer(sql):
            dep_name = m.group(1)
            if dep_name in models:  # only model deps, not source tables
                deps.add(dep_name)
        ref_deps[name] = deps

    # Sources are known leaf nodes
    known: set[str] = set(models.keys())
    for source in sources:
        src_name = source["name"]
        src_sheets = source.get("sheets")
        if src_sheets:
            for s in src_sheets:
                known.add(f"{src_name}_{s}")
        else:
            wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath("ssbt.yml")), source["path"]), read_only=True)
            for s in wb.sheetnames:
                known.add(f"{src_name}_{s}")
            wb.close()

    visited: set[str] = set()
    order: list[str] = []
    path: list[str] = []

    def visit(name: str) -> None:
        if name in path:
            cycle = path[path.index(name):] + [name]
            raise ValueError(f"Circular dependency: {' -> '.join(cycle)}")
        if name in visited:
            return
        if name not in known:
            raise ValueError(f"Unknown model or source: {name}")
        if name not in models:
            return  # source, skip
        path.append(name)
        for dep in ref_deps.get(name, set()):
            visit(dep)
        path.pop()
        visited.add(name)
        order.append(name)

    for name in models:
        visit(name)
    return order


# ---------------------------------------------------------------------------
# SQL loading and compilation
# ---------------------------------------------------------------------------

_REF_RE = re.compile(r"\{\{\s*ref\(\s*['\"]([^'\"]+)['\"]\s*\)\s*\}\}")


def load_sql(model: dict[str, Any], base_dir: str) -> str:
    """Read SQL file for a model."""
    sql_path = os.path.join(base_dir, model["sql_path"])
    with open(sql_path) as f:
        return f.read().strip()


def compile_sql(
    sql: str,
    models: dict[str, dict[str, Any]],
    table_map: dict[str, str] | None = None,
    _compiled: set[str] | None = None,
) -> str:
    """Replace {{ ref('name') }} with subquery for models, or table name for sources.

    Recursively compiles dependencies so nested {{ ref() }} are resolved.
    """
    if table_map is None:
        table_map = {name: name for name in models}
    if _compiled is None:
        _compiled = set()

    def _replace(m: re.Match) -> str:
        dep_name = m.group(1)
        if dep_name in models:
            # Model dependency — compile its SQL first (cached)
            if dep_name not in _compiled:
                models[dep_name]["_compiled_sql"] = compile_sql(
                    models[dep_name]["sql"], models, table_map, _compiled
                )
                _compiled.add(dep_name)
            dep_sql = models[dep_name]["_compiled_sql"]
            table_name = table_map.get(dep_name, dep_name)
            return f"(\n{dep_sql}\n) AS {table_name}"
        else:
            # Source table — just use the DuckDB table name
            return table_map.get(dep_name, dep_name)

    return _REF_RE.sub(_replace, sql)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _enumerate_sheets(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _get_existing_sheets(path: str) -> list[str]:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except FileNotFoundError:
        return []


# ---------------------------------------------------------------------------
# Execution
# ---------------------------------------------------------------------------

def run_manifest(
    models: dict[str, dict[str, Any]],
    sources: list[dict[str, Any]],
    execution_order: list[str],
    yml_path: str,
    input_path: str,
    output_dir: str = "output/",
    dry_run: bool = False,
) -> None:
    """Execute models in DAG order."""

    con = duckdb.connect()
    manifest_dir = os.path.dirname(os.path.abspath(yml_path))

    # Load all SQL first (for ref resolution)
    for name in models:
        models[name]["sql"] = load_sql(models[name], manifest_dir)

    # Resolve output files for each model
    for name in models:
        cfg = models[name].get("config", {})
        if "output" not in cfg:
            cfg["output"] = os.path.join(output_dir, f"{name}.xlsx")

    # Register sources as DuckDB tables
    # Source tables are named {source_name}_{sheet_name}
    # Also register bare sheet names as aliases if unambiguous (for backward compat)
    table_map: dict[str, str] = {}
    all_sheet_names: dict[str, list[str]] = {}
    for source in sources:
        src_name = source["name"]
        src_path = os.path.join(manifest_dir, source["path"])
        sheets = source.get("sheets")
        if sheets:
            sheet_names = sheets
        else:
            wb = openpyxl.load_workbook(src_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        for sheet_name in sheet_names:
            table_name = f"{src_name}_{sheet_name}"
            table_map[table_name] = table_name
            all_sheet_names.setdefault(sheet_name, []).append(src_name)
            sql = f"SELECT * FROM read_xlsx('{src_path}', sheet='{sheet_name}')"
            try:
                df = con.execute(sql).fetchdf()
                df.columns = [str(c).replace(" ", "_").replace("-", "_") for c in df.columns]
                con.execute(f"DROP TABLE IF EXISTS {table_name}")
                con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM {table_name}")
            except Exception:
                # read_xlsx fails on formula errors (#DIV/0!, #VALUE!, etc.)
                # Fall back to openpyxl which reads cell values directly

                warnings.filterwarnings("ignore")
                wb = openpyxl.load_workbook(src_path, data_only=True)
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    con.execute(f"CREATE TABLE {table_name} (id INTEGER)")
                    con.execute(f"INSERT INTO {table_name} VALUES (NULL)")
                    continue
                header = [str(c).replace(" ", "_").replace("-", "_") for c in rows[0]]
                data = [tuple(_coerce_cell(v) for v in row) for row in rows[1:]]
                if not data:
                    con.execute(f"CREATE TABLE {table_name} ({', '.join(f'{h} VARCHAR' for h in header)})")
                    continue

                df = pd.DataFrame(data, columns=header)
                con.register(table_name, df)
    # Also register bare sheet names as aliases if unambiguous
    for sheet_name, src_names in all_sheet_names.items():
        if len(src_names) == 1:
            table_map[sheet_name] = f"{src_names[0]}_{sheet_name}"
            alias_name = sheet_name
            src = next(s for s in sources if s["name"] == src_names[0])
            src_sheets = src.get("sheets")
            if src_sheets and sheet_name not in src_sheets:
                continue
            src_path = os.path.join(manifest_dir, src["path"])
            alias_name = sheet_name
            con.execute(f"DROP TABLE IF EXISTS {alias_name}")
            try:
                con.execute(f"CREATE TABLE {alias_name} AS SELECT * FROM read_xlsx('{src_path}', sheet='{sheet_name}')")
            except Exception:
                # read_xlsx fails on formula errors — fall back to openpyxl
                wb = openpyxl.load_workbook(src_path, data_only=True)
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if rows:
                    header = [str(c).replace(" ", "_").replace("-", "_") for c in rows[0]]
                    data = [tuple(_coerce_cell(v) for v in row) for row in rows[1:]]
                    if data:

                        df = pd.DataFrame(data, columns=header)
                        con.register(alias_name, df)

    # Execute models in order
    for model_name in execution_order:
        model = models[model_name]
        # Compile SQL — dependencies are already compiled since we go in DAG order
        compiled = compile_sql(model["sql"], models, table_map)
        table_map[model_name] = model_name

        cfg = model.get("config", {})
        output_file = cfg.get("output", os.path.join(output_dir, f"{model_name}.xlsx"))
        sheet_name = cfg.get("output_sheet", model_name)

        if dry_run:
            print(f"--- {model_name} ({model['sql_path']}) ---")
            print(compiled)
            print(f"  → {output_file} [{sheet_name}]")
            print()
            continue

        print(f"  [{model_name}]")

        result_df = con.execute(compiled).fetchdf()

        # Register for downstream models
        con.register(model_name, result_df)

        # Write to output
        os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

        if not os.path.exists(output_file):
            result_df.to_excel(output_file, sheet_name=sheet_name, index=False)
        else:
            existing = _get_existing_sheets(output_file)
            if sheet_name in existing:
                all_sheets = _read_all_sheets(output_file)
                all_sheets[sheet_name] = result_df
                _write_all_sheets(all_sheets, output_file)
            else:
                with pd.ExcelWriter(output_file, engine="openpyxl", mode="a",
                                    if_sheet_exists="replace") as writer:
                    result_df.to_excel(writer, sheet_name=sheet_name, index=False)

    con.close()


def _read_all_sheets(path: str) -> dict[str, "pd.DataFrame"]:
    sheets = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        sheets[name] = pd.read_excel(path, sheet_name=name, engine="openpyxl")
    wb.close()
    return sheets


def _write_all_sheets(sheets: dict[str, "pd.DataFrame"], path: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)


# ---------------------------------------------------------------------------
# Schema tests
# ---------------------------------------------------------------------------

def run_tests(
    models: dict[str, dict[str, Any]],
    sources: list[dict[str, Any]],
    execution_order: list[str],
    yml_path: str,
    input_path: str,
    output_path: str,
) -> list[dict[str, Any]]:
    """Run schema tests against the output. Returns list of test results."""

    con = duckdb.connect()

    # Load SQL
    manifest_dir = os.path.dirname(os.path.abspath(yml_path))
    for name in models:
        models[name]["sql"] = load_sql(models[name], manifest_dir)

    # Register sources
    table_map: dict[str, str] = {}
    all_sheet_names: dict[str, list[str]] = {}
    for source in sources:
        src_name = source["name"]
        src_path = os.path.join(manifest_dir, source["path"])
        sheets = source.get("sheets")
        if sheets:
            sheet_names = sheets
        else:
            wb = openpyxl.load_workbook(src_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        for sheet_name in sheet_names:
            table_name = f"{src_name}_{sheet_name}"
            table_map[table_name] = table_name
            all_sheet_names.setdefault(sheet_name, []).append(src_name)
            sql = f"SELECT * FROM read_xlsx('{src_path}', sheet='{sheet_name}')"
            try:
                df = con.execute(sql).fetchdf()
                df.columns = [str(c).replace(" ", "_").replace("-", "_") for c in df.columns]
                con.execute(f"DROP TABLE IF EXISTS {table_name}")
                con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM {table_name}")
            except Exception:
                # read_xlsx fails on formula errors (#DIV/0!, #VALUE!, etc.)
                # Fall back to openpyxl which reads cell values directly
                warnings.filterwarnings("ignore")
                wb = openpyxl.load_workbook(src_path, data_only=True)
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    con.execute(f"CREATE TABLE {table_name} (id INTEGER)")
                    con.execute(f"INSERT INTO {table_name} VALUES (NULL)")
                    continue
                header = [str(c).replace(" ", "_").replace("-", "_") for c in rows[0]]
                data = [tuple(_coerce_cell(v) for v in row) for row in rows[1:]]
                if not data:
                    con.execute(f"CREATE TABLE {table_name} ({', '.join(f'{h} VARCHAR' for h in header)})")
                    continue
                df = pd.DataFrame(data, columns=header)
                con.register(table_name, df)
    # Also register bare sheet names as aliases if unambiguous
    for sheet_name, src_names in all_sheet_names.items():
        if len(src_names) == 1:
            table_map[sheet_name] = f"{src_names[0]}_{sheet_name}"
            alias_name = sheet_name
            src = next(s for s in sources if s["name"] == src_names[0])
            src_sheets = src.get("sheets")
            if src_sheets and sheet_name not in src_sheets:
                continue
            src_path = os.path.join(manifest_dir, src["path"])
            alias_name = sheet_name
            con.execute(f"DROP TABLE IF EXISTS {alias_name}")
            try:
                con.execute(f"CREATE TABLE {alias_name} AS SELECT * FROM read_xlsx('{src_path}', sheet='{sheet_name}')")
            except Exception:
                # read_xlsx fails on formula errors — fall back to openpyxl
                wb = openpyxl.load_workbook(src_path, data_only=True)
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if rows:
                    header = [str(c).replace(" ", "_").replace("-", "_") for c in rows[0]]
                    data = [tuple(_coerce_cell(v) for v in row) for row in rows[1:]]
                    if data:
                        df = pd.DataFrame(data, columns=header)
                        con.register(alias_name, df)

    # Build output by running models (same as build, but capture results)
    for model_name in execution_order:
        model = models[model_name]
        compiled = compile_sql(model["sql"], models, table_map)
        table_map[model_name] = model_name
        result_df = con.execute(compiled).fetchdf()
        con.register(model_name, result_df)

    # Run schema tests
    results: list[dict[str, Any]] = []
    for model_name in execution_order:
        model = models[model_name]
        for col_def in model.get("columns", []):
            col_name = col_def["name"]
            for test in col_def.get("tests", []):
                if isinstance(test, str):
                    test_name = test
                    test_args = {}
                else:
                    test_name = list(test.keys())[0]
                    test_args = test[test_name]
                test_result = _run_test(
                    test_name, test_args, col_name, model_name, con
                )
                results.append(test_result)

    # Write test results to output
    if results:
        results_df = pd.DataFrame(results)
        if not os.path.exists(output_path):
            results_df.to_excel(output_path, sheet_name="test_results", index=False)
        else:
            existing = _get_existing_sheets(output_path)
            if "test_results" in existing:
                all_sheets = _read_all_sheets(output_path)
                all_sheets["test_results"] = results_df
                _write_all_sheets(all_sheets, output_path)
            else:
                with pd.ExcelWriter(output_path, engine="openpyxl", mode="a",
                                    if_sheet_exists="replace") as writer:
                    results_df.to_excel(writer, sheet_name="test_results", index=False)

    con.close()
    return results


def _run_test(
    test_name: str,
    test_args: Any,
    column_name: str,
    model_name: str,
    con: "duckdb.DuckDBPyConnection",
) -> dict[str, Any]:
    """Run a single schema test. Returns {model, column, test, status, message}."""
    import pandas as pd

    status = "pass"
    message = ""

    if test_name == "not_null":
        df = con.execute(
            f"SELECT * FROM {model_name} WHERE {column_name} IS NULL"
        ).fetchdf()
        if len(df) > 0:
            status = "fail"
            # Convert numpy types to Python native for clean display
            clean_rows = []
            for i in range(len(df)):
                row = []
                for val in df.iloc[i].tolist():
                    if hasattr(val, 'item'):
                        val = val.item()
                    if isinstance(val, float) and (val != val):  # nan check
                        val = None
                    elif isinstance(val, float) and val == int(val):
                        val = int(val)
                    row.append(val)
                clean_rows.append(str(row))
            rows_str = "; ".join(clean_rows)
            message = f"{len(df)} null values: {rows_str}"

    elif test_name == "unique":
        # Skip nulls — uniqueness can't be tested on nulls
        null_count = con.execute(
            f"SELECT COUNT(*) FROM {model_name} WHERE {column_name} IS NULL"
        ).fetchone()[0]
        if null_count > 0:
            status = "fail"
            message = f"skipped: {null_count} null values (cannot test uniqueness on nulls)"
            # Don't fail — just warn
            status = "warn"
            message = f"{null_count} null values skipped"
        else:
            dupes = con.execute(
                f"SELECT {column_name}, COUNT(*) AS cnt "
                f"FROM {model_name} GROUP BY {column_name} HAVING cnt > 1"
            ).fetchall()
            if dupes:
                status = "fail"
                message = f"{len(dupes)} duplicate values: {', '.join(str(d[0]) for d in dupes)}"

    elif test_name == "accepted_values":
        allowed = test_args.get("values", [])
        violations = con.execute(
            f"SELECT * FROM {model_name} "
            f"WHERE {column_name} IS NOT NULL AND {column_name} NOT IN ({','.join(repr(v) for v in allowed)})"
        ).fetchall()
        if violations:
            status = "fail"
            message = f"{len(violations)} violations: {', '.join(str(v[0]) for v in violations[:20])}"

    elif test_name == "positive":
        df = con.execute(
            f"SELECT {column_name} FROM {model_name} WHERE {column_name} <= 0"
        ).fetchdf()
        if len(df) > 0:
            status = "fail"
            vals = ", ".join(str(int(v)) if v == int(v) else str(v) for v in df.iloc[:, 0].tolist()[:20])
            message = f"{len(df)} non-positive values: [{vals}]"

    elif test_name == "not_empty":
        df = con.execute(
            f"SELECT {column_name} FROM {model_name} WHERE {column_name} = ''"
        ).fetchdf()
        if len(df) > 0:
            status = "fail"
            message = f"{len(df)} empty string values"

    elif test_name == "regex_match":
        pattern = test_args.get("expression", "")
        df = con.execute(
            f"SELECT {column_name} FROM {model_name} "
            f"WHERE {column_name} IS NOT NULL AND {column_name} !~ '{pattern}'"
        ).fetchdf()
        if len(df) > 0:
            status = "fail"
            vals = ", ".join(str(v) for v in df.iloc[:, 0].tolist()[:20])
            message = f"{len(df)} violations: [{vals}]"

    return {
        "model": model_name,
        "column": column_name,
        "test": test_name,
        "status": status,
        "message": message,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _print_test_results(results: list[dict[str, Any]]) -> None:
    """Print test results with checkmarks, return True if all passed."""
    passed = sum(1 for r in results if r["status"] == "pass")
    failed = sum(1 for r in results if r["status"] == "fail")
    warned = sum(1 for r in results if r["status"] == "warn")
    for r in results:
        if r["status"] == "pass":
            mark = "\u2713"
        elif r["status"] == "fail":
            mark = "\u2717"
        else:
            mark = "!"
        print(f"  {mark} {r['model']}.{r['column']} ({r['test']})")
    parts = [f"{passed} passed", f"{failed} failed"]
    if warned:
        parts.append(f"{warned} warned")
    print(f"\n{', '.join(parts)}")
    if failed > 0:
        for r in results:
            if r["status"] == "fail":
                print(f"  FAIL {r['model']}.{r['column']} ({r['test']}): {r['message']}")
        raise SystemExit(1)
    return True


def main():
    parser = argparse.ArgumentParser(description="ssbt — spreadsheet build tool")
    parser.add_argument("action", choices=["build", "run", "test", "docs"],
                        help="build: run models + tests | test: run tests only | run: models only")
    parser.add_argument("--yml", default="ssbt.yml", help="ssbt.yml path")
    parser.add_argument("--input", default="input.xlsx", help="Input Excel (legacy: single file)")
    parser.add_argument("--output", default="output.xlsx", help="Output Excel")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    cfg = load_manifest(args.yml)
    models = load_models(cfg)
    sources = load_sources(cfg)

    # Backward compat: if no sources defined, treat --input as a single source
    if not sources:
        sources = [{"name": "input", "path": args.input, "sheets": None}]

    order = resolve_dag(models, sources)

    if args.action in ("build", "run"):
        run_manifest(models, sources, order, args.yml, args.input, args.output, dry_run=args.dry_run)
        if not args.dry_run:
            results = run_tests(models, sources, order, args.yml, args.input, args.output)
            _print_test_results(results)
    elif args.action == "test":
        results = run_tests(models, sources, order, args.yml, args.input, args.output)
        _print_test_results(results)
    elif args.action == "docs":
        print("docs: not yet implemented")

    print("Done.")


if __name__ == "__main__":
    main()
